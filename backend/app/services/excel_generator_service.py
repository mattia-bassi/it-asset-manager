from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart3D, Reference
import openpyxl.chart.label
from pathlib import Path
from typing import List, Dict, Optional
from datetime import datetime

import logging

logger = logging.getLogger(__name__)


class ExcelGeneratorService:
    """Servizio per generare file Excel con branding aziendale"""
    
    # Colori aziendali GBSAPRI
    COLOR_YELLOW = "FFDD0F"
    COLOR_GRAY = "515D64"
    COLOR_WHITE = "FFFFFF"
    COLOR_LIGHT_GRAY = "E8E8E8"
    
    BASE_PATH = Path("/app")
    OUTPUT_DIR = BASE_PATH / "data" / "reports"
    
    @classmethod
    def _setup_workbook(cls, wb: Workbook, title: str, logo_path: Optional[str] = None):
        """Setup iniziale del workbook con logo e header"""
        ws = wb.active
        ws.title = "Report"
        
        # Imposta larghezza colonne A-E per l'area logo (25 = ~187px per colonna)
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 25
        
        # Logo (se presente) - GRANDE: dalla A alla E, righe 1-5
        current_row = 1
        if logo_path and Path(logo_path).exists():
            try:
                img = XLImage(logo_path)
                # Scala logo per occupare da A1 a E5
                # Larghezza: circa 5 colonne = ~650px per rimanere dentro A-E
                img.width = 650
                # Mantieni proporzioni
                img.height = int(img.height * (650 / img.width))
                # Limita altezza massima a 5 righe (~150px)
                if img.height > 150:
                    img.height = 150
                    img.width = int(img.width * (150 / img.height))
                # Merge celle A1:E5 per il logo
                ws.merge_cells('A1:E5')
                logo_cell = ws['A1']
                logo_cell.alignment = Alignment(horizontal='center', vertical='center')
                ws.add_image(img, 'A1')
                # Aumenta altezza righe per il logo (1-5)
                for row_num in range(1, 6):
                    ws.row_dimensions[row_num].height = 30
                current_row = 6  # Logo occupa righe 1-5
            except Exception as e:
                logger.warning("Errore caricamento logo: %s", e)
                current_row = 1
        else:
            current_row = 1
        
        # Titolo report (dopo il logo)
        ws.merge_cells(f'A{current_row}:E{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = title
        title_cell.font = Font(name='Arial', size=16, bold=True, color=cls.COLOR_GRAY)
        title_cell.fill = PatternFill(start_color=cls.COLOR_YELLOW, end_color=cls.COLOR_YELLOW, fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 30
        
        # Info generazione
        current_row += 1
        ws[f'A{current_row}'] = f"Data generazione: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws[f'A{current_row}'].font = Font(name='Arial', size=9, italic=True)
        
        return ws, current_row + 2  # Ritorna riga dove iniziare dati
    
    @classmethod
    def _create_header_row(cls, ws, headers: List[str], start_row: int, start_col: int = 1):
        """Crea riga header con stile brand"""
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx, header in enumerate(headers, start=start_col):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = header
            cell.font = Font(name='Arial', size=11, bold=True, color=cls.COLOR_WHITE)
            cell.fill = PatternFill(start_color=cls.COLOR_GRAY, end_color=cls.COLOR_GRAY, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        ws.row_dimensions[start_row].height = 25
    
    @classmethod
    def _add_data_rows(cls, ws, data: List[Dict], start_row: int, headers: List[str]):
        """Aggiunge righe dati con alternanza colori"""
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_idx, row_data in enumerate(data, start=start_row):
            # Alterna colori righe
            fill_color = cls.COLOR_WHITE if row_idx % 2 == 0 else cls.COLOR_LIGHT_GRAY
            
            for col_idx, header_key in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Converti chiave header in chiave dati (es: "Tipo" -> "tipo")
                data_key = header_key.lower().replace(' ', '_')
                cell.value = row_data.get(data_key, '')
                
                cell.font = Font(name='Arial', size=10)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = border
        
        # Auto-size colonne
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

    @classmethod
    def _add_pie_chart_3d(cls, ws, title: str, data_start_row: int, data_end_row: int,
                          label_col: int = 1, data_col: int = 2, anchor_cell: str = None):
        """Aggiunge grafico a torta 3D al foglio"""
        if data_end_row <= data_start_row:
            return  # Nessun dato

        chart = PieChart3D()
        chart.title = title
        chart.style = 10
        chart.width = 26   # largo per dare spazio
        chart.height = 18  # alto per separare torta e leggenda

        # Leggenda sotto il grafico (non sovrapposta alla torta)
        chart.legend.position = 'b'

        # Dati (colonna valori)
        data_ref = Reference(ws, min_col=data_col, min_row=data_start_row, max_row=data_end_row)
        # Etichette (colonna nomi)
        labels_ref = Reference(ws, min_col=label_col, min_row=data_start_row + 1, max_row=data_end_row)

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(labels_ref)

        # Mostra percentuali sulle fette
        chart.dataLabels = openpyxl.chart.label.DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showCatName = False
        chart.dataLabels.showVal = False

        # Posiziona il grafico a destra, partendo da riga 4
        if anchor_cell is None:
            anchor_cell = "H1"
        ws.add_chart(chart, anchor_cell)

    @classmethod
    def generate_assets_by_type_report(
        cls,
        data: List[Dict],
        logo_path: Optional[str] = None
    ) -> str:
        """Report 1: Asset per Tipo"""
        cls.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        
        wb = Workbook()
        ws, start_row = cls._setup_workbook(wb, "REPORT ASSET PER TIPO", logo_path)
        
        headers = ['Tipo', 'Totale', 'Assegnati', 'Disponibili', 'Guasti']
        cls._create_header_row(ws, headers, start_row)
        cls._add_data_rows(ws, data, start_row + 1, headers)

        # Grafico a torta 3D - Distribuzione asset per tipo
        if data:
            data_end_row = start_row + len(data)
            cls._add_pie_chart_3d(ws, "Distribuzione Asset per Tipo", start_row, data_end_row)

        filename = f"report_asset_per_tipo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = cls.OUTPUT_DIR / filename
        wb.save(str(filepath))
        
        return f"/data/reports/{filename}"
    
    @classmethod
    def generate_faulty_assets_report(
        cls,
        data: List[Dict],
        logo_path: Optional[str] = None
    ) -> str:
        """Report 2: Dispositivi Guasti"""
        cls.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        
        wb = Workbook()
        ws, start_row = cls._setup_workbook(wb, "REPORT DISPOSITIVI GUASTI", logo_path)
        
        headers = ['Marca', 'Modello', 'Tipo', 'Numero Guasti']
        cls._create_header_row(ws, headers, start_row)
        cls._add_data_rows(ws, data, start_row + 1, headers)

        if data:
            data_end_row = start_row + len(data)
            cls._add_pie_chart_3d(ws, "Dispositivi Guasti per Tipo", start_row, data_end_row,
                                  label_col=3, data_col=4)

        filename = f"report_dispositivi_guasti_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = cls.OUTPUT_DIR / filename
        wb.save(str(filepath))
        
        return f"/data/reports/{filename}"
    
    @classmethod
    def generate_active_assignments_report(
        cls,
        data: List[Dict],
        logo_path: Optional[str] = None
    ) -> str:
        """Report 3: Assegnazioni Attive"""
        cls.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        
        wb = Workbook()
        ws, start_row = cls._setup_workbook(wb, "REPORT ASSEGNAZIONI ATTIVE", logo_path)
        
        headers = ['Numero Assegnazione', 'Data', 'Persona', 'Email', 'Sede', 'Num Items']
        cls._create_header_row(ws, headers, start_row)
        cls._add_data_rows(ws, data, start_row + 1, headers)

        if data:
            data_end_row = start_row + len(data)
            cls._add_pie_chart_3d(ws, "Assegnazioni Attive per Sede", start_row, data_end_row,
                                  label_col=5, data_col=6)

        filename = f"report_assegnazioni_attive_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = cls.OUTPUT_DIR / filename
        wb.save(str(filepath))
        
        return f"/data/reports/{filename}"
    
    @classmethod
    def generate_assignment_history_report(
        cls,
        data: List[Dict],
        logo_path: Optional[str] = None
    ) -> str:
        """Report 4: Storico Assegnazioni"""
        cls.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        
        wb = Workbook()
        ws, start_row = cls._setup_workbook(wb, "REPORT STORICO ASSEGNAZIONI", logo_path)
        
        headers = ['Numero Assegnazione', 'Data Assegnazione', 'Data Riconsegna', 'Durata Giorni', 'Stato', 'Persona', 'Sede', 'Num Items']
        cls._create_header_row(ws, headers, start_row)
        cls._add_data_rows(ws, data, start_row + 1, headers)
        
        filename = f"report_storico_assegnazioni_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = cls.OUTPUT_DIR / filename
        wb.save(str(filepath))
        
        return f"/data/reports/{filename}"
    
    @classmethod
    def generate_low_stock_report(
        cls,
        data: List[Dict],
        logo_path: Optional[str] = None
    ) -> str:
        """Report 5: Inventario Sotto Soglia"""
        cls.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        
        wb = Workbook()
        ws, start_row = cls._setup_workbook(wb, "REPORT INVENTARIO SOTTO SOGLIA", logo_path)
        
        headers = ['Categoria', 'Dispositivo', 'Marca', 'Quantità Attuale', 'Quantità Minima', 'Percentuale', 'Stato Alert']
        cls._create_header_row(ws, headers, start_row)
        cls._add_data_rows(ws, data, start_row + 1, headers)

        if data:
            data_end_row = start_row + len(data)
            cls._add_pie_chart_3d(ws, "Materiali Sotto Soglia", start_row, data_end_row,
                                  label_col=1, data_col=4)

        filename = f"report_inventario_sotto_soglia_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = cls.OUTPUT_DIR / filename
        wb.save(str(filepath))
        
        return f"/data/reports/{filename}"
    
    @classmethod
    def generate_assets_by_site_report(
        cls,
        data: List[Dict],
        logo_path: Optional[str] = None
    ) -> str:
        """Report 6: Asset per Sede"""
        cls.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        
        wb = Workbook()
        ws, start_row = cls._setup_workbook(wb, "REPORT ASSET PER SEDE", logo_path)
        
        headers = ['Sede', 'Tipo', 'Totale']
        cls._create_header_row(ws, headers, start_row)
        cls._add_data_rows(ws, data, start_row + 1, headers)

        if data:
            data_end_row = start_row + len(data)
            cls._add_pie_chart_3d(ws, "Distribuzione Asset per Sede", start_row, data_end_row)

        filename = f"report_asset_per_sede_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = cls.OUTPUT_DIR / filename
        wb.save(str(filepath))
        
        return f"/data/reports/{filename}"
    
    @classmethod
    def generate_report(
        cls,
        title: str,
        data: List[Dict],
        logo_path: Optional[str] = None
    ) -> str:
        """
        Genera un report Excel generico da una lista di dizionari
        
        Args:
            title: Titolo del report
            data: Lista di dizionari con i dati (le chiavi diventano header)
            logo_path: Percorso opzionale del logo
            
        Returns:
            Percorso del file Excel generato
        """
        if not data:
            # Se non ci sono dati, crea comunque il file con messaggio
            data = [{"Messaggio": "Nessun dato disponibile"}]
        
        # Estrai headers dalle chiavi del primo elemento
        headers = list(data[0].keys())
        
        # Crea workbook
        wb = Workbook()
        ws, start_row = cls._setup_workbook(wb, title, logo_path)
        
        # Aggiungi header
        cls._create_header_row(ws, headers, start_row)
        
        # Aggiungi dati (usa direttamente le chiavi senza conversione)
        cls._add_data_rows_generic(ws, data, start_row + 1, headers)
        
        # Salva file
        cls.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"report_{timestamp}.xlsx"
        filepath = cls.OUTPUT_DIR / filename
        
        wb.save(str(filepath))
        
        return str(filepath)
    
    @classmethod
    def _add_data_rows_generic(cls, ws, data: List[Dict], start_row: int, headers: List[str]):
        """Aggiunge righe dati con alternanza colori (versione generica che usa chiavi esatte)"""
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_idx, row_data in enumerate(data, start=start_row):
            # Alterna colori righe
            fill_color = cls.COLOR_WHITE if row_idx % 2 == 0 else cls.COLOR_LIGHT_GRAY
            
            for col_idx, header_key in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Usa direttamente la chiave dell'header (senza conversione)
                cell.value = row_data.get(header_key, '')
                
                cell.font = Font(name='Arial', size=10)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = border
        
        # Auto-size colonne
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
